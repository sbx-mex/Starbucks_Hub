#!/usr/bin/env python3
"""Construye el dashboard estático desde Forms + Directorio Centro Norte.

Fuentes editoriales:
  - cms/Sistema de Evidencias OPS.xlsx
  - cms/Centro Norte_Directorio.xlsx
  - cms/Sistema_Evidencias_OPS_CMS.xlsx

El navegador nunca procesa los Excel. Este motor valida encabezados, cruza CeCo,
deduplica cumplimiento por tienda/actividad y publica únicamente el JSON mínimo.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_RESPONSES = ROOT / "cms" / "Sistema de Evidencias OPS.xlsx"
DEFAULT_DIRECTORY = ROOT / "cms" / "Centro Norte_Directorio.xlsx"
DEFAULT_ACTIVITIES = ROOT / "config" / "actividades.csv"
DEFAULT_MANAGERS = ROOT / "config" / "gerentes.csv"
DEFAULT_CMS = ROOT / "cms" / "Sistema_Evidencias_OPS_CMS.xlsx"
DEFAULT_SETTINGS = ROOT / "config" / "settings.json"
DEFAULT_OUTPUT = ROOT / "data" / "dashboard.json"

RESPONSE_FIELDS = {
    "id": ("Id",),
    "started": ("Hora de inicio",),
    "finished": ("Hora de finalización", "Hora de finalizacion"),
    "email": ("Correo electrónico", "Correo electronico"),
    "name": ("Nombre",),
    "activity": ("Selecciona la actividad que deseas registrar",),
    "ceco": ("CeCo",),
    "confirmed": ("¿Confirmas que realizaste la actividad seleccionada?",),
    "evidence": ("Evidencia del avance",),
}


def clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def key_text(value: Any) -> str:
    text = unicodedata.normalize("NFD", clean_text(value).casefold())
    return "".join(char for char in text if unicodedata.category(char) != "Mn")


def is_yes(value: Any) -> bool:
    return key_text(value) in {"si", "true", "1", "yes"}


def normalize_ceco(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        value = str(int(value))
    text = clean_text(value).replace(".0", "")
    digits = re.sub(r"\D", "", text)
    return digits if len(digits) == 5 else ""


def parse_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    text = clean_text(value)
    for fmt in ("%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def resolve_columns(headers: list[Any], contract: dict[str, tuple[str, ...]]) -> dict[str, int]:
    normalized = {key_text(value): index for index, value in enumerate(headers) if clean_text(value)}
    result: dict[str, int] = {}
    missing = []
    for field, aliases in contract.items():
        match = next((normalized[key_text(alias)] for alias in aliases if key_text(alias) in normalized), None)
        if match is None:
            missing.append(aliases[0])
        else:
            result[field] = match
    if missing:
        raise ValueError("Faltan encabezados requeridos: " + ", ".join(missing))
    return result


def load_settings(path: Path, cms_settings: dict[str, Any] | None = None) -> dict[str, Any]:
    settings = json.loads(path.read_text(encoding="utf-8"))
    settings.update(cms_settings or {})
    return settings


def parse_date(value: Any):
    if isinstance(value, datetime):
        return value.date()
    text = clean_text(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Fecha CMS inválida: {text}")


def date_status(start, end) -> str:
    today = datetime.now().date()
    if start and today < start:
        return "Programada"
    if end and today > end:
        return "Vencida"
    return "Vigente"


def find_header(ws, required: set[str]) -> tuple[int, dict[str, int]]:
    for row_number, row in enumerate(ws.iter_rows(min_row=1, max_row=12, values_only=True), 1):
        normalized = {key_text(value): index for index, value in enumerate(row) if clean_text(value)}
        if required.issubset(normalized):
            return row_number, normalized
    raise ValueError(f"No se encontró encabezado {', '.join(sorted(required))} en {ws.title}")


def load_cms(path: Path) -> tuple[list[dict[str, Any]], dict[str, dict[str, str]], dict[str, Any], dict[str, int]]:
    """Lee actividades, fechas, gerentes y configuración desde un solo Excel CMS."""
    workbook = load_workbook(path, read_only=True, data_only=False)
    required_sheets = {"Actividades", "Gerentes", "Configuracion"}
    missing = required_sheets.difference(workbook.sheetnames)
    if missing:
        raise ValueError("Faltan hojas CMS: " + ", ".join(sorted(missing)))

    config_ws = workbook["Configuracion"]
    config_header, config_cols = find_header(config_ws, {"clave", "valor"})
    cms_settings: dict[str, Any] = {}
    boolean_keys = {"onlyOpenStores", "requireEvidence", "publishEvidenceLinks", "publishPersonalData"}
    for row in config_ws.iter_rows(min_row=config_header + 1, values_only=True):
        key = clean_text(row[config_cols["clave"]])
        if not key:
            continue
        value = row[config_cols["valor"]]
        cms_settings[key] = is_yes(value) if key in boolean_keys else clean_text(value)

    activity_ws = workbook["Actividades"]
    header_row, cols = find_header(activity_ws, {"orden", "actividad", "descripcion", "fecha inicio", "fecha limite", "activo"})
    activities = []
    calendar = {"active": 0, "scheduled": 0, "expired": 0, "inactive": 0}
    for row in activity_ws.iter_rows(min_row=header_row + 1, values_only=True):
        name = clean_text(row[cols["actividad"]])
        if not name:
            continue
        active = is_yes(row[cols["activo"]])
        start = parse_date(row[cols["fecha inicio"]])
        end = parse_date(row[cols["fecha limite"]])
        if start and end and end < start:
            raise ValueError(f"La fecha límite de {name} es anterior a la fecha de inicio")
        status = date_status(start, end)
        if not active:
            calendar["inactive"] += 1
            continue
        if status == "Programada":
            calendar["scheduled"] += 1
            continue
        if status == "Vencida":
            calendar["expired"] += 1
            continue
        calendar["active"] += 1
        evidence_col = cols.get("evidencia requerida")
        priority_col = cols.get("prioridad")
        activities.append({
            "name": name,
            "description": clean_text(row[cols["descripcion"]]),
            "order": int(float(row[cols["orden"]] or 999)),
            "startDate": start.isoformat() if start else None,
            "endDate": end.isoformat() if end else None,
            "commitmentDateDisplay": end.strftime("%d/%m/%y") if end else "Sin fecha compromiso",
            "requireEvidence": is_yes(row[evidence_col]) if evidence_col is not None else True,
            "priority": clean_text(row[priority_col]) if priority_col is not None else "Media",
            "autoDetected": False,
        })

    manager_ws = workbook["Gerentes"]
    manager_header, manager_cols = find_header(manager_ws, {"dm", "nombre corto", "foto webp", "activo"})
    managers: dict[str, dict[str, str]] = {}
    for row in manager_ws.iter_rows(min_row=manager_header + 1, values_only=True):
        dm = clean_text(row[manager_cols["dm"]])
        if not dm or not is_yes(row[manager_cols["activo"]]):
            continue
        photo = clean_text(row[manager_cols["foto webp"]])
        if photo and not (ROOT / photo).is_file():
            raise ValueError(f"No existe la fotografía configurada para {dm}: {photo}")
        managers[key_text(dm)] = {
            "shortName": clean_text(row[manager_cols["nombre corto"]]) or dm,
            "photo": photo,
        }
    return sorted(activities, key=lambda item: (item["order"], key_text(item["name"]))), managers, cms_settings, calendar


def load_activities(path: Path) -> list[dict[str, Any]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    activities = []
    seen = set()
    for row in rows:
        name = clean_text(row.get("Actividad"))
        if not name or not is_yes(row.get("Activo")) or key_text(name) in seen:
            continue
        seen.add(key_text(name))
        activities.append({
            "name": name,
            "description": clean_text(row.get("Descripción")),
            "order": int(float(row.get("Orden") or 999)),
            "autoDetected": False,
        })
    return sorted(activities, key=lambda item: (item["order"], key_text(item["name"])))


def load_managers(path: Path) -> dict[str, dict[str, str]]:
    """Carga el catálogo visual de DM y valida que sus fotografías existan."""
    with path.open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    managers: dict[str, dict[str, str]] = {}
    for row in rows:
        dm = clean_text(row.get("DM"))
        if not dm or not is_yes(row.get("Activo")):
            continue
        photo = clean_text(row.get("Foto"))
        if photo and not (ROOT / photo).is_file():
            raise ValueError(f"No existe la fotografía configurada para {dm}: {photo}")
        managers[key_text(dm)] = {
            "shortName": clean_text(row.get("Nombre corto")) or dm,
            "photo": photo,
        }
    return managers


def status_label(compliance: float) -> str:
    if compliance >= 80:
        return "En meta"
    if compliance >= 40:
        return "Seguimiento"
    return "Atención"


def find_directory_header(ws) -> tuple[int, list[Any]]:
    for row_number, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 12), values_only=True), 1):
        keys = {key_text(value) for value in row if clean_text(value)}
        if {"cc", "cc nombre", "dm"}.issubset(keys):
            return row_number, list(row)
    raise ValueError(f"No se encontró encabezado CC / CC Nombre / DM en {ws.title}")


def load_directory(path: Path, settings: dict[str, Any]) -> tuple[dict[str, dict[str, str]], str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    requested = settings.get("directorySheet")
    if requested not in workbook.sheetnames:
        requested = max(workbook.sheetnames, key=lambda name: workbook[name].max_row)
    ws = workbook[requested]
    header_row, headers = find_directory_header(ws)
    normalized = {key_text(value): index for index, value in enumerate(headers) if clean_text(value)}
    required = ("cc", "cc nombre", "region", "estatus", "dm")
    missing = [field for field in required if field not in normalized]
    if missing:
        raise ValueError("Directorio incompleto: " + ", ".join(missing))

    stores: dict[str, dict[str, str]] = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        ceco = normalize_ceco(row[normalized["cc"]])
        if not ceco:
            continue
        region = clean_text(row[normalized["region"]])
        status = clean_text(row[normalized["estatus"]])
        if settings.get("region") and key_text(region) != key_text(settings["region"]):
            continue
        if settings.get("onlyOpenStores") and key_text(status) not in {"abierta", "abierto", "activa", "activo"}:
            continue
        stores[ceco] = {
            "ceco": ceco,
            "store": clean_text(row[normalized["cc nombre"]]) or f"Tienda {ceco}",
            "dm": clean_text(row[normalized["dm"]]) or "Sin DM",
            "region": region,
            "status": status,
        }
    return stores, requested


def load_responses(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    ws = workbook[workbook.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows, ()))
    columns = resolve_columns(headers, RESPONSE_FIELDS)
    responses = []
    for row_number, row in enumerate(rows, 2):
        if not any(value not in (None, "") for value in row):
            continue
        finished = parse_datetime(row[columns["finished"]])
        responses.append({
            "row": row_number,
            "id": clean_text(row[columns["id"]]) or str(row_number - 1),
            "started": parse_datetime(row[columns["started"]]),
            "finished": finished,
            "email": clean_text(row[columns["email"]]),
            "name": clean_text(row[columns["name"]]),
            "activity": clean_text(row[columns["activity"]]),
            "ceco": normalize_ceco(row[columns["ceco"]]),
            "confirmed": is_yes(row[columns["confirmed"]]),
            "evidence": clean_text(row[columns["evidence"]]),
        })
    return responses


def iso_or_none(value: datetime | None) -> str | None:
    return value.isoformat(timespec="seconds") if value else None


def build_payload(
    responses_path: Path,
    directory_path: Path,
    settings_path: Path,
    cms_path: Path = DEFAULT_CMS,
) -> dict[str, Any]:
    activities, managers, cms_settings, calendar = load_cms(cms_path)
    settings = load_settings(settings_path, cms_settings)
    stores, directory_sheet = load_directory(directory_path, settings)
    responses = load_responses(responses_path)

    configured = {key_text(item["name"]): item for item in activities}
    for response in responses:
        activity_key = key_text(response["activity"])
        if activity_key and activity_key not in configured:
            item = {
                "name": response["activity"],
                "description": "Actividad detectada automáticamente en las respuestas del Forms.",
                "order": 900 + len(configured),
                "startDate": None,
                "endDate": None,
                "commitmentDateDisplay": "Sin fecha compromiso",
                "requireEvidence": settings.get("requireEvidence", True),
                "priority": "Media",
                "autoDetected": True,
            }
            activities.append(item)
            configured[activity_key] = item
    activities.sort(key=lambda item: (item["order"], key_text(item["name"])))
    activity_names = [item["name"] for item in activities]
    canonical_activity = {key_text(item["name"]): item["name"] for item in activities}
    evidence_rules = {key_text(item["name"]): item.get("requireEvidence", True) for item in activities}

    submissions = []
    latest_by_pair: dict[tuple[str, str], dict[str, Any]] = {}
    unknown_cecos = set()
    invalid_rows = []
    latest_update = None

    for response in responses:
        store = stores.get(response["ceco"])
        activity = canonical_activity.get(key_text(response["activity"]), response["activity"])
        evidence_available = bool(re.match(r"^https?://", response["evidence"], flags=re.I))
        valid = bool(
            store
            and activity
            and response["confirmed"]
            and (evidence_available or not evidence_rules.get(key_text(activity), settings.get("requireEvidence", True)))
        )
        if response["ceco"] and not store:
            unknown_cecos.add(response["ceco"])
        if not valid:
            invalid_rows.append(response["row"])
        if response["finished"] and (latest_update is None or response["finished"] > latest_update):
            latest_update = response["finished"]

        public = {
            "id": response["id"],
            "timestamp": iso_or_none(response["finished"]),
            "timestampDisplay": response["finished"].strftime("%d/%m/%Y %H:%M") if response["finished"] else "Sin fecha",
            "activity": activity or "Sin actividad",
            "ceco": response["ceco"] or "Inválido",
            "store": store["store"] if store else "CeCo sin cruce",
            "dm": store["dm"] if store else "Sin asignar",
            "confirmed": response["confirmed"],
            "evidenceAvailable": evidence_available,
            "valid": valid,
        }
        if settings.get("publishEvidenceLinks") and evidence_available:
            public["evidenceUrl"] = response["evidence"]
        if settings.get("publishPersonalData"):
            public["submittedBy"] = response["name"]
            public["email"] = response["email"]
        submissions.append(public)

        if valid:
            pair = (response["ceco"], activity)
            current = latest_by_pair.get(pair)
            if current is None or (response["finished"] or datetime.min) > (current["finished"] or datetime.min):
                latest_by_pair[pair] = response

    completion_pairs = set(latest_by_pair)
    store_rows = []
    for ceco, store in sorted(stores.items(), key=lambda item: (key_text(item[1]["dm"]), key_text(item[1]["store"]))):
        status = {activity: (ceco, activity) in completion_pairs for activity in activity_names}
        completed = sum(status.values())
        timestamps = [item["finished"] for (pair_ceco, _), item in latest_by_pair.items() if pair_ceco == ceco and item["finished"]]
        store_rows.append({
            **store,
            "completed": completed,
            "expected": len(activity_names),
            "compliance": round(completed / len(activity_names) * 100, 1) if activity_names else 0,
            "lastUpdate": iso_or_none(max(timestamps)) if timestamps else None,
            "activities": status,
        })

    activity_stats = []
    for item in activities:
        completed = sum((ceco, item["name"]) in completion_pairs for ceco in stores)
        activity_stats.append({
            **item,
            "completedStores": completed,
            "pendingStores": len(stores) - completed,
            "compliance": round(completed / len(stores) * 100, 1) if stores else 0,
        })

    dm_groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for store in store_rows:
        dm_groups[store["dm"]].append(store)
    dm_stats = []
    for dm, dm_stores in sorted(dm_groups.items(), key=lambda item: key_text(item[0])):
        completed = sum(store["completed"] for store in dm_stores)
        expected = sum(store["expected"] for store in dm_stores)
        compliance = round(completed / expected * 100, 1) if expected else 0
        profile = managers.get(key_text(dm), {})
        pending_stores = sum(store["completed"] < store["expected"] for store in dm_stores)
        dm_stats.append({
            "dm": dm,
            "shortName": profile.get("shortName", dm),
            "photo": profile.get("photo", ""),
            "stores": len(dm_stores),
            "completed": completed,
            "expected": expected,
            "pending": expected - completed,
            "pendingStores": pending_stores,
            "compliance": compliance,
            "status": status_label(compliance),
        })
    dm_stats.sort(key=lambda item: (-item["compliance"], key_text(item["shortName"])))
    for rank, item in enumerate(dm_stats, 1):
        item["rank"] = rank

    expected_total = len(stores) * len(activity_names)
    completed_total = len(completion_pairs)
    valid_responses = sum(item["valid"] for item in submissions)
    stores_complete = sum(item["completed"] == item["expected"] and item["expected"] > 0 for item in store_rows)

    return {
        "schemaVersion": 4,
        "project": settings.get("projectName", "Sistema de Evidencias OPS"),
        "region": settings.get("region", "Centro Norte"),
        "generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
        "lastUpdated": iso_or_none(latest_update),
        "lastUpdatedDisplay": latest_update.strftime("%d/%m/%Y %H:%M") if latest_update else "Sin respuestas",
        "report": {
            "title": "Sistema de Evidencia OPS",
            "subtitle": "Dashboard de Avance de Actividades",
            "motto": "JUNTÉMONOS MÁS",
            "credits": "Diseñado por Jorge Alcantar Aguiar & Enrique César Flores",
            "cutOffDisplay": latest_update.strftime("%d/%m/%y · %H:%M h") if latest_update else "Sin datos",
        },
        "sources": {
            "responses": responses_path.name,
            "directory": directory_path.name,
            "directorySheet": directory_sheet,
            "cms": cms_path.name,
        },
        "summary": {
            "dms": len(dm_stats),
            "stores": len(stores),
            "activities": len(activity_names),
            "expectedCompletions": expected_total,
            "completedCompletions": completed_total,
            "compliance": round(completed_total / expected_total * 100, 1) if expected_total else 0,
            "validResponses": valid_responses,
            "storesComplete": stores_complete,
            "pendingCompletions": expected_total - completed_total,
        },
        "quality": {
            "responsesRead": len(responses),
            "invalidRows": invalid_rows,
            "unknownCeCos": sorted(unknown_cecos),
            "duplicateValidResponses": max(valid_responses - completed_total, 0),
            "privacyMode": not settings.get("publishPersonalData") and not settings.get("publishEvidenceLinks"),
        },
        "calendar": calendar,
        "activities": activity_stats,
        "dms": dm_stats,
        "attention": sorted(
            [
                {
                    "ceco": store["ceco"],
                    "store": store["store"],
                    "dm": store["dm"],
                    "completed": store["completed"],
                    "expected": store["expected"],
                    "pending": store["expected"] - store["completed"],
                    "compliance": store["compliance"],
                    "status": status_label(store["compliance"]),
                }
                for store in store_rows
                if store["completed"] < store["expected"]
            ],
            key=lambda item: (item["compliance"], key_text(item["dm"]), key_text(item["store"])),
        ),
        "stores": store_rows,
        "submissions": sorted(submissions, key=lambda item: item["timestamp"] or "", reverse=True),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Genera data/dashboard.json desde los Excel del proyecto.")
    parser.add_argument("--responses", type=Path, default=DEFAULT_RESPONSES)
    parser.add_argument("--directory", type=Path, default=DEFAULT_DIRECTORY)
    parser.add_argument("--settings", type=Path, default=DEFAULT_SETTINGS)
    parser.add_argument("--cms", type=Path, default=DEFAULT_CMS)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    payload = build_payload(args.responses, args.directory, args.settings, args.cms)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    summary = payload["summary"]
    print(
        f"Dashboard generado: {summary['stores']} tiendas · {summary['activities']} actividades · "
        f"{summary['completedCompletions']}/{summary['expectedCompletions']} cumplimientos"
    )
    print(f"Última actualización Forms: {payload['lastUpdatedDisplay']}")


if __name__ == "__main__":
    main()
